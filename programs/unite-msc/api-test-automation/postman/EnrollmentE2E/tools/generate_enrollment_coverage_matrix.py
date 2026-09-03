#!/usr/bin/env python3
"""Generate enrollment automation coverage matrix from Dinesh's Excel + api-test-automation repo."""

from __future__ import annotations

import re
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "Enrollment End Points.xlsx"
OUT_XLSX = ROOT / "Enrollment-Automation-Coverage-Matrix.xlsx"
OUT_MD = ROOT / "Enrollment-Automation-Coverage-Status.md"

# endpoint path (normalized key) -> automation metadata
AUTOMATION_MAP = {
    "/enrollmentapi/health/liveness": {
        "status": "Done",
        "test_class": "EnrollmentPingRequestTest",
        "method": "getEnrollmentLiveness",
        "suites": "smoke",
        "scope": "Infrastructure",
        "legacy_note": "Not in legacy Cucumber; added in MSC TestNG smoke",
        "remaining": "",
    },
    "/enrollmentapi/v1/ping": {
        "status": "Done",
        "test_class": "EnrollmentPingRequestTest",
        "method": "getEnrollmentPing_returnsOk",
        "suites": "smoke",
        "scope": "Core E2E — bootstrap",
        "legacy_note": "Migrated from legacy; old collection marked Done",
        "remaining": "",
    },
    "/enrollmentapi/v1/certificate": {
        "status": "Done",
        "test_class": "EnrollmentCertificateRequestTest",
        "suites": "smoke",
        "scope": "Core E2E — bootstrap",
        "legacy_note": "New in MSC automation (encryption setup)",
        "remaining": "",
    },
    "/enrollmentapi/v1/usstates": {
        "status": "Done",
        "test_class": "EnrollmentUsStatesRequestTest",
        "suites": "smoke",
        "scope": "Core E2E — bootstrap",
        "legacy_note": "New in MSC automation",
        "remaining": "",
    },
    "/enrollmentapi/v1/country": {
        "status": "Done",
        "test_class": "EnrollmentCountryRequestTest",
        "suites": "smoke",
        "scope": "Core E2E — bootstrap",
        "legacy_note": "New in MSC automation",
        "remaining": "",
    },
    "/enrollmentapi/v1/plans": {
        "status": "Done",
        "test_class": "EnrollmentPlansRequestTest",
        "method": "getEnrollmentPlans_returnsPlans",
        "suites": "smoke",
        "scope": "Core E2E — bootstrap",
        "legacy_note": "Migrated from legacy; old collection marked Done",
        "remaining": "",
    },
    "/enrollmentapi/v1/plans/{planId}": {
        "status": "Done",
        "test_class": "EnrollmentPlansRequestTest",
        "method": "getEnrollmentPlanById_returnsPlan",
        "suites": "smoke",
        "scope": "Core E2E — bootstrap",
        "legacy_note": "Migrated from legacy; old collection marked Done",
        "remaining": "",
    },
    "/enrollmentapi/v1/content?branding={planId}&language=en&name=enrollment": {
        "status": "Done",
        "test_class": "EnrollmentContentRequestTest",
        "suites": "regression, integration",
        "scope": "Core E2E — wizard",
        "legacy_note": "New in MSC automation (not in legacy collection)",
        "remaining": "",
    },
    "/mobile1api/v1/mobilemembersession": {
        "status": "Done",
        "test_class": "MobileMemberSessionRequestTest",
        "suites": "smoke; regression (okdirect only)",
        "scope": "Optional",
        "legacy_note": "New — cross-API (Mobile1); optional Postman step",
        "remaining": "",
    },
    "/enrollmentapi/v1/enrollments/enrollmentstarted": {
        "status": "Done",
        "test_class": "EnrollmentStartedRequestTest",
        "suites": "regression, integration",
        "scope": "Optional — web flow",
        "legacy_note": "New in MSC automation; web account-owner UI only",
        "remaining": "",
    },
    "/enrollmentapi/v1/enrollments/prospects": {
        "status": "Done",
        "test_class": "ProspectRequestTest",
        "suites": "regression, integration",
        "scope": "Core E2E — wizard",
        "legacy_note": "Migrated from legacy; old collection marked Done",
        "remaining": "",
    },
    "/enrollmentapi/v1/enrollments/enrollment/owner-entered": {
        "status": "Done",
        "test_class": "OwnerEnteredTests",
        "suites": "regression, integration",
        "scope": "Core E2E — wizard",
        "legacy_note": "Migrated from legacy; old collection marked Done",
        "remaining": "",
    },
    "/enrollmentapi/v1/enrollments/enrollment/owner-address-entered": {
        "status": "Done",
        "test_class": "OwnerAddressEnteredRequestTest",
        "suites": "regression, integration",
        "scope": "Core E2E — wizard",
        "legacy_note": "New in MSC automation (not listed in old collection)",
        "remaining": "",
    },
    "/enrollmentapi/v1/enrollments/enrollment/beneficiary-entered": {
        "status": "Done",
        "test_class": "BeneficiaryEnteredTests",
        "suites": "regression, integration",
        "scope": "Core E2E — wizard",
        "legacy_note": "Migrated from legacy; old collection marked Done",
        "remaining": "",
    },
    "/enrollmentapi/v1/verify/routingnumber": {
        "status": "Done",
        "test_class": "VerifyBankRoutingNumberRequestTest",
        "suites": "regression, integration",
        "scope": "Core E2E — wizard",
        "legacy_note": "MSC automation done; old collection still marked In Progress",
        "remaining": "Confirm Stage1 green + update Jira QA-1603",
    },
    "/enrollmentapi/v1/enrollments/enrollment/bank-entered": {
        "status": "Done",
        "test_class": "BankEnteredRequestTests",
        "suites": "regression, integration",
        "scope": "Core E2E — wizard",
        "legacy_note": "Migrated from legacy; old collection marked Done",
        "remaining": "",
    },
    "/enrollmentapi/v1/enrollments/enrollment/recurring-contribution-entered": {
        "status": "Done",
        "test_class": "RecurringContributionEnteredRequestTest",
        "suites": "regression, integration",
        "scope": "Optional — AIP",
        "legacy_note": "New in MSC automation; optional skip in Postman",
        "remaining": "",
    },
    "/enrollmentapi/v1/enrollmentallocationfunds/get": {
        "status": "Done",
        "test_class": "AllocationFundRequestTest",
        "suites": "regression, integration",
        "scope": "Optional — investment helper",
        "legacy_note": "New in MSC automation; alternative to SQL fund lookup",
        "remaining": "",
    },
    "/enrollmentapi/v1/enrollments/enrollment/allocations-entered": {
        "status": "Done",
        "test_class": "AllocationsEnteredRequestTests",
        "suites": "regression, integration",
        "scope": "Core E2E — wizard",
        "legacy_note": "MSC automation done; old collection marked In Progress",
        "remaining": "Confirm Stage1 green + update Jira QA-1602",
    },
    "/enrollmentapi/v1/enrollments/enrollment/review-confirm-entered": {
        "status": "Done",
        "test_class": "ReviewConfirmEnteredRequestTest",
        "suites": "regression, integration",
        "scope": "Core E2E — submit",
        "legacy_note": "New in MSC automation; QA-1604 checked in Sep 2026",
        "remaining": "Confirm Stage1 green on okdirect + newyork; nmdirect still localhost-only",
    },
    "/enrollmentapi/v1/subsequentenrollment/banks": {
        "status": "Done",
        "test_class": "SubsequentEnrollmentBanksRequestTest",
        "suites": "regression, integration",
        "scope": "Subsequent E2E",
        "legacy_note": "New in MSC; QA-1792 — not in legacy collection",
        "remaining": "Add nmdirect to CI suites",
    },
    "/enrollmentapi/v1/enrollments/subsequentenrollment/beneficiary-entered": {
        "status": "Done",
        "test_class": "SubsequentBeneficiaryEnteredRequestTest",
        "suites": "regression, integration",
        "scope": "Subsequent E2E",
        "legacy_note": "New in MSC; QA-1853 — not in Enrollment End Points.xlsx catalog",
        "remaining": "Add row to Dinesh Excel catalog",
    },
    "/enrollmentapi/v1/enrollments/subsequentenrollment/bank-entered": {
        "status": "Done",
        "test_class": "SubsequentEnrollmentBankEnteredRequestTest",
        "suites": "regression, integration",
        "scope": "Subsequent E2E",
        "legacy_note": "New in MSC; QA-1854 — not in Enrollment End Points.xlsx catalog",
        "remaining": "Add row to Dinesh Excel catalog",
    },
    "/enrollmentapi/v1/enrollments/subsequentenrollment/recurring-contribution-entered": {
        "status": "Done",
        "test_class": "SubsequentEnrollmentRecurringContributionRequestTest",
        "suites": "regression, integration",
        "scope": "Subsequent E2E",
        "legacy_note": "New in MSC; QA-1855 — not in Enrollment End Points.xlsx catalog",
        "remaining": "Add row to Dinesh Excel catalog",
    },
    "/enrollmentapi/v1/enrollments/subsequentenrollment/review-confirm-entered": {
        "status": "Done",
        "test_class": "SubsequentEnrollmentReviewConfirmEnteredRequestTest",
        "suites": "regression, integration",
        "scope": "Subsequent E2E",
        "legacy_note": "New in MSC; QA-1791 — not in legacy collection",
        "remaining": "Add nmdirect to CI suites",
    },
    "/enrollmentapi/v1/enrollments/submit": {
        "status": "Deferred",
        "test_class": "—",
        "suites": "—",
        "scope": "Out of Scope — partner (Vanguard)",
        "legacy_note": "Partner-only; Postman 401",
        "remaining": "QA-1808 — partner integration research",
    },
    "/enrollmentapi/v1/upromiseaccount": {
        "status": "Deferred",
        "test_class": "—",
        "suites": "—",
        "scope": "Out of Scope — partner",
        "legacy_note": "Partner-only; auth failed in Postman",
        "remaining": "QA-1807",
    },
    "/enrollmentapi/v1/oauth/token": {
        "status": "Deferred",
        "test_class": "—",
        "suites": "—",
        "scope": "Out of Scope — OAuth",
        "legacy_note": "Partner/service auth; Postman 401",
        "remaining": "Not planned for MSC E2E",
    },
}

HEADER_FILL = PatternFill("solid", fgColor="003241")
DONE_FILL = PatternFill("solid", fgColor="D1FAE5")
PROGRESS_FILL = PatternFill("solid", fgColor="FEF3C7")
NOT_FILL = PatternFill("solid", fgColor="FEE2E2")
DEFER_FILL = PatternFill("solid", fgColor="E2E8F0")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)


def normalize_endpoint(ep: str) -> str:
    return ep.strip()


def classify_legacy(old_status: str | None) -> str:
    if not old_status:
        return "New in MSC (not in legacy collection)"
    s = str(old_status).strip()
    if "not assigned" in s.lower() or "not started" in s.lower():
        return "New in MSC (legacy: not assigned)"
    if "done" in s.lower():
        return "Migrated from legacy (marked Done)"
    if "in progress" in s.lower() or "inprogress" in s.lower():
        return "Legacy existed — MSC automation ahead of old status"
    return f"Legacy: {s}"


def status_fill(status: str) -> PatternFill:
    if status == "Done":
        return DONE_FILL
    if status == "Not Started":
        return NOT_FILL
    if status == "Deferred":
        return DEFER_FILL
    return PROGRESS_FILL


def read_source_rows() -> list[dict]:
    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb["API Endpoints - Enrollment"]
    rows = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r[2] or not str(r[2]).startswith("/"):
            continue
        rows.append(
            {
                "category": r[0] or "",
                "method": r[1] or "",
                "endpoint": normalize_endpoint(str(r[2])),
                "postman_status": r[3] or "",
                "old_collection_status": r[4] or "",
                "auth_level": r[5] or "",
                "description": r[6] or "",
                "use_case": r[7] or "",
            }
        )
    return rows


def program_scope_for(scope: str) -> str:
    if scope.startswith("Core E2E") or scope.startswith("Subsequent"):
        return "In Scope"
    if scope.startswith("Optional") or scope == "Infrastructure":
        return "In Scope (optional)"
    return "Out of Scope (deferred)"


EXTRA_CATALOG_ROWS = [
    {
        "category": "Subsequent Enrollment (Java only)",
        "method": "POST",
        "endpoint": "/enrollmentapi/v1/enrollments/subsequentenrollment/beneficiary-entered",
        "postman_status": "In Postman E2E as POST 25",
        "old_collection_status": "",
        "auth_level": "Member JWT",
        "description": "Subsequent beneficiary-entered",
        "use_case": "Add another 529 account for existing member",
    },
    {
        "category": "Subsequent Enrollment (Java only)",
        "method": "POST",
        "endpoint": "/enrollmentapi/v1/enrollments/subsequentenrollment/bank-entered",
        "postman_status": "In Postman E2E as POST 26",
        "old_collection_status": "",
        "auth_level": "Member JWT",
        "description": "Subsequent bank-entered",
        "use_case": "Add another 529 account for existing member",
    },
    {
        "category": "Subsequent Enrollment (Java only)",
        "method": "POST",
        "endpoint": "/enrollmentapi/v1/enrollments/subsequentenrollment/recurring-contribution-entered",
        "postman_status": "In Postman E2E as POST 27",
        "old_collection_status": "",
        "auth_level": "Member JWT",
        "description": "Subsequent recurring contribution",
        "use_case": "Add another 529 account for existing member",
    },
]


def build_matrix_rows(source_rows: list[dict]) -> list[dict]:
    out = []
    seen = {normalize_endpoint(str(r["endpoint"])) for r in source_rows}
    combined = list(source_rows)
    for extra in EXTRA_CATALOG_ROWS:
        if extra["endpoint"] not in seen:
            combined.append(extra)
    for i, row in enumerate(combined, start=1):
        ep = row["endpoint"]
        auto = AUTOMATION_MAP.get(ep, {})
        scope = auto.get("scope", "TBD")
        program_scope = program_scope_for(scope)

        out.append(
            {
                "#": i,
                **row,
                "automation_status": auto.get("status", "TBD"),
                "java_test_class": auto.get("test_class", "—"),
                "testng_suites": auto.get("suites", "—"),
                "program_scope": program_scope,
                "scope_detail": scope,
                "legacy_migration": classify_legacy(row["old_collection_status"]),
                "msc_automation_note": auto.get("legacy_note", ""),
                "remaining_work": auto.get("remaining", ""),
            }
        )
    return out


def write_excel(rows: list[dict]) -> None:
    wb = openpyxl.Workbook()

    # Tab 1 — Coverage matrix
    ws = wb.active
    ws.title = "Automation Coverage"

    headers = [
        "#",
        "Category",
        "Method",
        "Endpoint",
        "Postman Status",
        "Old Collection Status",
        "Automation Status",
        "Java Test Class",
        "TestNG Suites",
        "Program Scope",
        "Scope Detail",
        "Legacy / Migration",
        "MSC Automation Notes",
        "Remaining Work",
        "Auth Level",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = WHITE_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in rows:
        ws.append(
            [
                r["#"],
                r["category"],
                r["method"],
                r["endpoint"],
                r["postman_status"],
                r["old_collection_status"],
                r["automation_status"],
                r["java_test_class"],
                r["testng_suites"],
                r["program_scope"],
                r["scope_detail"],
                r["legacy_migration"],
                r["msc_automation_note"],
                r["remaining_work"],
                r["auth_level"],
            ]
        )
        row_idx = ws.max_row
        fill = status_fill(r["automation_status"])
        ws.cell(row=row_idx, column=7).fill = fill

    widths = [4, 16, 8, 52, 14, 22, 14, 34, 22, 18, 22, 28, 36, 30, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # Tab 2 — Summary
    ws2 = wb.create_sheet("Summary")
    done = sum(1 for r in rows if r["automation_status"] == "Done")
    not_started = sum(1 for r in rows if r["automation_status"] == "Not Started")
    deferred = sum(1 for r in rows if r["automation_status"] == "Deferred")
    core = [r for r in rows if r["scope_detail"].startswith("Core E2E")]
    core_done = sum(1 for r in core if r["automation_status"] == "Done")
    optional_done = sum(
        1 for r in rows if r["program_scope"] == "In Scope (optional)" and r["automation_status"] == "Done"
    )

    summary = [
        ("Enrollment Automation Coverage Summary", ""),
        ("Generated", str(date.today())),
        ("Source mapping", "Enrollment End Points.xlsx (Dinesh)"),
        ("Repo", "api-test-automation/mobile/enrollment"),
        ("", ""),
        ("Total endpoints (catalog)", len(rows)),
        ("Java automated (Done)", done),
        ("Not started (in scope)", not_started),
        ("Deferred (out of scope)", deferred),
        ("Core E2E endpoints", f"{core_done}/{len(core)}"),
        ("Optional helpers automated", optional_done),
        ("", ""),
        ("Blocking gap", "None for happy-path wizard + subsequent (okdirect + newyork)"),
        ("Remaining", "nmdirect CI; negatives; partner submit/Upromise/OAuth"),
    ]
    for a, b in summary:
        ws2.append([a, b])
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 50
    ws2["A1"].font = BOLD

    # Tab 3 — App flow sequence (from source tab 2)
    ws3 = wb.create_sheet("E2E Flow Sequence")
    ws3.append(["Step", "Phase", "Endpoint / Action", "Notes"])
    for col in range(1, 5):
        ws3.cell(row=1, column=col).fill = HEADER_FILL
        ws3.cell(row=1, column=col).font = WHITE_FONT

    src_wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    flow_ws = src_wb["Enrollment App Flow"]
    for r in flow_ws.iter_rows(min_row=4, values_only=True):
        if not any(r):
            continue
        step, phase, arrow, endpoint, note = (list(r) + [None] * 5)[:5]
        if endpoint and ("GET" in str(endpoint) or "POST" in str(endpoint)):
            ep_match = re.search(r"(/\S+)", str(endpoint))
            ep_key = None
            if ep_match:
                raw = ep_match.group(1).rstrip(",")
                for k in AUTOMATION_MAP:
                    if raw.startswith(k.split("?")[0].replace("{planId}", "").rstrip("/")):
                        ep_key = k
                        break
            auto = AUTOMATION_MAP.get(ep_key, {}) if ep_key else {}
            note_text = auto.get("status", "") or (note or "")
            ws3.append([step or "", phase or "", str(endpoint).strip(), note_text])
    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 70
    ws3.column_dimensions["D"].width = 28

    wb.save(OUT_XLSX)


def write_markdown(rows: list[dict]) -> None:
    done = [r for r in rows if r["automation_status"] == "Done"]
    pending = [r for r in rows if r["automation_status"] == "Not Started"]
    deferred = [r for r in rows if r["automation_status"] == "Deferred"]
    core = [r for r in rows if r["scope_detail"].startswith("Core E2E")]
    core_done = [r for r in core if r["automation_status"] == "Done"]

    lines = [
        "# Enrollment API — Automation Coverage Status",
        "",
        f"**Generated:** {date.today()}  ",
        "**Source of truth:** `Enrollment End Points.xlsx` (tab: API Endpoints - Enrollment)  ",
        "**Repo:** `api-test-automation/mobile/enrollment`  ",
        "**Matrix Excel:** [Enrollment-Automation-Coverage-Matrix.xlsx](./Enrollment-Automation-Coverage-Matrix.xlsx)",
        "",
        "---",
        "",
        "## Summary",
        "",
        "| Metric | Count |",
        "|--------|-------|",
        f"| Total endpoints in catalog | **{len(rows)}** |",
        f"| **Java automated (Done)** | **{len(done)}** ({100*len(done)//len(rows)}%) |",
        f"| Core E2E wizard + submit | **{len(core_done)}/{len(core)}** ({100*len(core_done)//len(core) if core else 0}%) |",
        f"| Not started (in scope) | **{len(pending)}** |",
        f"| Deferred (out of scope) | **{len(deferred)}** |",
        "",
        "**Coding status (Sep 2026):** Initial wizard including `review-confirm-entered` is Done. Subsequent banks / beneficiary / bank-entered / recurring / review-confirm are Done for **okdirect + newyork**. Remaining is documentation, CI plants (nmdirect), negatives, partner APIs.",
        "",
        "---",
        "",
        "## Legend",
        "",
        "| Automation Status | Meaning |",
        "|-------------------|---------|",
        "| **Done** | TestNG class exists in `mobile/enrollment` |",
        "| **Not Started** | In scope for current sprint; no Java class |",
        "| **Deferred** | Out of scope — partner submit / Upromise / OAuth |",
        "",
        "| Program Scope | Meaning |",
        "|---------------|---------|",
        "| In Scope | Required for MSC E2E happy path |",
        "| In Scope (optional) | In Postman flow; can skip (mobile login, AIP, allocation funds GET) |",
        "| Out of Scope (deferred) | Not targeting this sprint |",
        "",
        "| Legacy / Migration | Meaning |",
        "|--------------------|---------|",
        "| Migrated from legacy | Was in old Cucumber/Postman collection (marked Done) |",
        "| New in MSC | Added in new TestNG framework; not in legacy collection |",
        "| Legacy existed — MSC ahead | Old collection still \"In Progress\" but Java is Done |",
        "",
        "---",
        "",
        f"## Done — Java automated ({len(done)} endpoints)",
        "",
        "| # | Endpoint | Test Class | Suites | Legacy note |",
        "|---|----------|------------|--------|-------------|",
    ]
    for r in done:
        lines.append(
            f"| {r['#']} | `{r['endpoint'][:55]}` | `{r['java_test_class']}` | {r['testng_suites']} | {r['msc_automation_note'][:50]} |"
        )

    lines += [
        "",
        "---",
        "",
        "## Not started — in scope",
        "",
    ]
    if pending:
        lines.append("| # | Endpoint | Remaining work |")
        lines.append("|---|----------|----------------|")
        for r in pending:
            lines.append(f"| {r['#']} | `{r['endpoint']}` | {r['remaining_work']} |")
    else:
        lines.append("_None._")

    lines += [
        "",
        "---",
        "",
        "## Deferred — out of scope (next sprint / research)",
        "",
        "| # | Endpoint | Postman | Reason |",
        "|---|----------|---------|--------|",
    ]
    for r in deferred:
        lines.append(
            f"| {r['#']} | `{r['endpoint'][:50]}` | {r['postman_status']} | {r['remaining_work'] or r['scope_detail']} |"
        )

    lines += [
        "",
        "---",
        "",
        "## TestNG suite wiring",
        "",
        "| Suite | What runs |",
        "|-------|-----------|",
        "| `enrollment-smoke-testng.xml` | Bootstrap GETs + optional mobile login (okdirect) |",
        "| `enrollment-regression-testng.xml` | Full wizard + subsequent (okdirect + newyork) |",
        "| `enrollment-integration-testng.xml` | Same as regression on QC4 |",
        "| `localhost-testng.xml.example` | Local three-plan shell including nmdirect — not CI |",
        "",
        "**Note:** `MobileMemberSessionRequestTest` is `groups=functional` so it is listed in regression XML but filtered out of regression/integration runs.",
        "",
        "---",
        "",
        "## New in MSC vs migrated from legacy",
        "",
        "### Migrated from legacy (old collection marked Done)",
        "",
    ]
    for r in rows:
        if "Migrated from legacy" in r["legacy_migration"]:
            lines.append(f"- `{r['endpoint']}` → `{r['java_test_class']}`")

    lines += ["", "### New in MSC automation (not in legacy collection)", ""]
    for r in rows:
        if "New in MSC" in r["legacy_migration"] and r["automation_status"] == "Done":
            lines.append(f"- `{r['endpoint']}` → `{r['java_test_class']}`")

    lines += [
        "",
        "---",
        "",
        "## Regenerate",
        "",
        "```powershell",
        "cd programs/unite-msc/api-test-automation/postman/EnrollmentE2E/tools",
        "python generate_enrollment_coverage_matrix.py",
        "```",
        "",
    ]
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    source = read_source_rows()
    rows = build_matrix_rows(source)
    write_excel(rows)
    write_markdown(rows)
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_MD}")
    print(f"Endpoints: {len(rows)} | Done: {sum(1 for r in rows if r['automation_status']=='Done')}")


if __name__ == "__main__":
    main()
