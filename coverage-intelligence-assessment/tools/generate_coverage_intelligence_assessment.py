#!/usr/bin/env python3
"""Generate coverage-intelligence-assessment CSV/XLSX artifacts from audit evidence."""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
ASSESSMENT_DATE = "2026-07-20T16:45:00-04:00"


def write_csv(rel_path: str, headers: list[str], rows: list[list]) -> None:
    path = ROOT / rel_path
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)


def build_integration_matrix() -> None:
    rows = [
        ["Atlassian Jira/Confluence MCP", "Cursor MCP user-jira", "Yes (config present)", "MCP session", "No", "N/A", "Tool discovery error; mcp_auth timeout", "Unknown", "Stale/unknown", "Jira REST API after credential fix", "Blocked", "GetMcpTools 2026-07-20; mcp_auth timeout"],
        ["qTest MCP", "Cursor MCP", "No", "N/A", "No", "N/A", "No qTest MCP server registered", "N/A", "N/A", "qTest REST API v3", "Not configured", "MCP catalog 2026-07-20 — no qTest server"],
        ["qTest REST API", "HTTPS + bearer token", "No", "API token (env)", "No", "N/A", "QTEST_BASE_URL, QTEST_API_TOKEN, QTEST_PROJECT_ID not set", "Unknown", "Unknown", "Scheduled Python collector", "Not configured", "Env scan 2026-07-20"],
        ["GitLab MCP", "Cursor MCP user-gitlab", "Yes", "MCP session", "No", "N/A", "Server error at discovery", "Unknown", "Unknown", "GitLab REST API / glab CLI", "Blocked", "GetMcpTools 2026-07-20"],
        ["GitLab REST API", "HTTPS PRIVATE-TOKEN", "Partial", "GITLAB_TOKEN env", "No", "N/A", "401 Unauthorized on /api/v4/user", "Rate limits per GitLab.com", "Unknown", "glab or REST after valid token", "Blocked", "Invoke-RestMethod 2026-07-20"],
        ["GitHub MCP", "Cursor MCP user-git", "Yes", "MCP session", "No", "N/A", "Server error at discovery", "Unknown", "Unknown", "gh CLI if configured", "Blocked", "GetMcpTools 2026-07-20"],
        ["GitHub CLI", "gh", "No", "N/A", "No", "N/A", "gh not installed", "N/A", "N/A", "Install gh + token for Actions evidence", "Not configured", "Shell which gh 2026-07-20"],
        ["Jenkins API", "HTTP/CLI", "No", "N/A", "No", "N/A", "No API URL or token in env", "N/A", "Unknown", "Exported job configs / console logs", "Partial", "KB exports in universal-platform-coverage/00-input-evidence"],
        ["SonarQube", "HTTP API", "No", "SONAR_TOKEN not set", "No", "N/A", "RUN_SONARQUBE: false on UniteMSC CI", "N/A", "Unknown", "Sonar API after access", "Not configured", "unite-mobile1/.gitlab-ci.yml; env scan"],
        ["Nexus", "HTTP", "No", "NEXUS_URL not set", "No", "N/A", "Referenced in POMs only", "N/A", "Unknown", "Artifact metadata export", "Not configured", "unite-mobile2/pom.xml distributionManagement"],
        ["Local qTest/Jenkins exports", "Filesystem", "Yes", "N/A", "Yes (read-only)", "qTest PDF 744 cases; Jenkins logs", "Row-level qTest IDs not exported", "N/A", "2026-06-29 qTest snapshot", "Reuse as historical evidence", "Partial", "universal-platform-coverage/00-input-evidence/v2-qtest-jenkins/"],
        ["Python coverage utilities", "Local scripts", "Yes", "N/A", "Yes", "UP rebuild; GS deliverables generator", "Read-only", "N/A", "2026-07-20", "Extend existing Python collectors", "Available", "universal-platform-coverage/03-tools/; government-savings-automation-assessment/tools/"],
        ["GS coverage KB (prior assessment)", "Filesystem", "Yes", "N/A", "Yes", "Repo/pipeline inventories", "No live system query", "N/A", "2026-07-20", "Baseline for this audit", "Available", "government-savings-automation-assessment/"],
        ["Brave Search MCP", "Cursor MCP", "Yes", "MCP", "Yes", "Web search", "Public web only", "API limits", "Live", "Supplemental research", "Available", "GetMcpTools ready"],
        ["Snyk MCP", "Cursor MCP", "Yes", "MCP", "Yes", "Security scans", "Not coverage scope", "N/A", "Live", "Dependency security only", "Available", "GetMcpTools ready"],
        ["1Password MCP", "Cursor MCP", "Yes", "MCP", "No", "N/A", "Server error at discovery", "N/A", "Unknown", "Credential retrieval if fixed", "Blocked", "GetMcpTools 2026-07-20"],
        ["Slack MCP", "Cursor MCP", "Yes", "MCP", "No", "N/A", "Server error", "N/A", "Unknown", "Notifications only", "Blocked", "GetMcpTools 2026-07-20"],
    ]
    write_csv(
        "01-connectivity/integration-capability-matrix.csv",
        ["System", "Access method", "Configured", "Authentication type", "Read access verified", "Available objects", "Permission limitations", "Rate or pagination limitations", "Data freshness", "Recommended extraction method", "Status", "Evidence"],
        rows,
    )


def build_qtest_data_quality() -> None:
    rows = [
        ["DQ-001", "Duplicate test cases", "Unknown", "Cannot verify without live qTest API", "High", "Export full test-case list with PID + name hash", "Blocked"],
        ["DQ-002", "Obsolete/inactive tests", "Inferred", "744 qTest population includes executed PRIME cycle; obsolete filter unknown", "Medium", "Filter by last execution + automation flag", "Stale export only"],
        ["DQ-003", "Tests without requirement linkage", "Inferred", "Module mapping provisional; Jira links not in export", "High", "qTest requirement link field audit", "Blocked"],
        ["DQ-004", "Automated without repo mapping", "Inferred", "qTest marks automation via PRIME cycle + TR-* run IDs in reports", "Medium", "Crosswalk qTest automation field to suite XML", "Partial"],
        ["DQ-005", "Repo tests without qTest mapping", "Verified", "Mobile2 API tests have no qTest IDs in source @ cee0de9", "High", "Add stable automation IDs in code + qTest", "Code scan"],
        ["DQ-006", "Manual counted as automation", "Inferred", "UP assessment excluded manual cases from automation %", "Medium", "Enforce automation status field in qTest", "Methodology"],
        ["DQ-007", "Not executed recently", "Stale", "qTest export dated 2026-06-29; freshness >21 days", "High", "Refresh qTest execution export", "Stale"],
        ["DQ-008", "Multiple versions same scenario", "Inferred", "179-case PRIME expansion unmapped — possible param variants", "Medium", "SME review of PRIME expansion bucket", "UP ledger"],
        ["DQ-009", "Missing business-area classification", "Verified", "Module-level only; 179 UNMAPPED_NEEDS_SME_CONFIRMATION", "High", "Custom field or tag standard", "v2-row-level-mapping.csv"],
        ["DQ-010", "Missing endpoint/service identifiers", "Verified", "qTest V2 UI cases not endpoint-keyed", "High", "Not applicable for UI; use for API catalog separately", "Scope distinction"],
        ["DQ-011", "Stale Jira links", "Unknown", "Cannot verify without Jira API", "High", "Jira-qTest link export", "Blocked"],
        ["DQ-012", "Stale owners", "Unknown", "Owner field not in available export", "Medium", "qTest owner + modified date report", "Blocked"],
        ["DQ-013", "Unclear regression designation", "Inferred", "PRIME cycle used as nightly regression proxy", "Medium", "Define approved regression suite IDs in qTest", "UP methodology"],
    ]
    write_csv(
        "02-qtest/qtest-data-quality-summary.csv",
        ["finding_id", "category", "status", "evidence", "severity", "remediation", "evidence_source"],
        rows,
    )


def build_jira_link_quality() -> None:
    rows = [
        ["ACS-I-2679", "Enrollment", "Inferred", "qTest module map", "PROVISIONAL-MOD-01", "Module-level only", "No row-level qTest ID", "Stale", "Refresh qTest export + Jira link API"],
        ["ACS-I-2680", "IDP Login", "Inferred", "qTest module map", "PROVISIONAL-MOD-02", "Module-level only", "No row-level qTest ID", "Stale", "Same"],
        ["ACS-I-2681", "ABLE", "Inferred", "qTest module map", "PROVISIONAL-MOD-04", "Module-level only", "No row-level qTest ID", "Stale", "Same"],
        ["ACS-I-2690", "Withdrawal", "Inferred", "qTest module map", "PROVISIONAL-MOD-03", "Module-level only", "No row-level qTest ID", "Stale", "Same"],
        ["QA-1405", "Mobile 2 GitLab nightly", "Inferred", "Jira story in KB", "N/A", "Pipeline spec only", "No qTest link verified", "Planned", "Link story to qTest cycle on implementation"],
        ["QA-1313", "Mobile 1 owner/profile", "Verified", "Git commit message", "N/A", "Code only", "No qTest link in commit", "Verified", "Add traceability on close"],
    ]
    write_csv(
        "03-jira/jira-qtest-link-quality.csv",
        ["jira_key", "domain", "link_status", "jira_evidence", "qtest_reference", "match_method", "gap", "freshness", "remediation"],
        rows,
    )


def build_repository_inventory() -> None:
    rows = [
        ["api-test-automation", r"C:\Workspace\GitLab\api-test-automation", "https://gitlab.com/ascensus-gs/products/depot/qa-automation/api-test-automation.git", "main", "cee0de9", "2026-07-20", "Active", "GS API MSC Universal ASTRO enrollment", "Maven Java TestNG RestAssured", "API integration regression", "25+ testng XML mobile2; metadataweb profiles", "stage1-api-metadata-pipeline; mobile profiles", "JUnit Surefire", "Sparse qTest refs in docs", "Endpoint paths in *RequestTest", "banks-smoke excluded from master", "DB properties per env", "Stage1 QC4", "IDP+non-IDP partial"],
        ["prime-test-automation", r"C:\Workspace\GitLab\prime-test-automation", "https://gitlab.com/ascensus-gs/products/depot/qa-automation/prime-test-automation.git", "main", "93f8628", "2026-07-06", "Active", "Unite V3 UI UE", "Maven Java Cucumber Selenium", "UI E2E", "stage1-ue-regression-test; stage1-unite-regression-master", "GitLab scheduled_regression_job", "JUnit", "Indirect via UP ledger", "N/A", "None identified", "Stage1 DB secure files", "Stage1 Stage5 smoke", "IDP enabled"],
        ["unite-test-automation", r"C:\Workspace\GitLab\Automation\unite-test-automation", "https://gitlab.com/ascensus-gs/products/depot/qa-automation/automation.git", "main", "14e3b62a", "2026-07-16", "Active legacy", "Unite V2 UI backoffice", "Ant TestNG Cucumber", "UI E2E batch", "stage1-*-regression suite XMLs", "Jenkins Ant targets", "TestNG HTML", "TR-* style in module docs", "Feature files", "Unknown quarantine", "TB refresh dependency", "Stage1 Stage5", "Partial IDP"],
        ["astro-test-automation", r"C:\Workspace\GitLab\Automation\astro-test-automation", "automation.git subpath", "main", "14e3b62a", "2026-07-16", "Active legacy", "ASTRO SFRP", "Ant TestNG Cucumber", "UI E2E", "astro testsuite", "Jenkins ASTRO-TB-REFRESH", "TestNG", "Sparse", "N/A", "Unknown", "TB1", "Partial"],
        ["performance-test-automation", r"C:\Workspace\GitLab\Automation\performance-test-automation", "automation.git subpath", "main", "14e3b62a", "2026-07-16", "Active", "GS performance", "JMeter Taurus YAML", "Performance", "49 JMX", "Jenkins AGSUP_* jobs", "Perf metrics", "Planned QA-S-PERF-005", "YAML service names", "N/A", "Load servers", "QC4 Stage1", "IDP/non-IDP variants"],
        ["unite-mobile1", r"C:\Workspace\GitLab\MobileAutomation\UniteMSC\unite-mobile1", "https://gitlab.com/ascensus-gs/products/unitemsc/unite-mobile1.git", "main", "bfb51aa", "2026-03-12", "Active", "MSC microservice", "Maven Java JUnit Cucumber", "Service BDD unit", "Cucumber features", "unitemsc_template.yml", "Jacoco Cluecumber", "N/A", "Service endpoints", "N/A", "Service CI", "Dev/test", "N/A"],
        ["unite-mobile2", r"C:\Workspace\GitLab\MobileAutomation\UniteMSC\unite-mobile2", "unitemsc/unite-mobile2.git", "main", "varies", "2026", "Active", "MSC BFF service", "Maven Java", "Service tests", "Cucumber", "unitemsc_template.yml", "Jacoco", "N/A", "BFF routes", "N/A", "Service CI", "Dev/test", "N/A"],
        ["qa-automation-kb", r"C:\Workspace\GitLab\qa-automation-kb", "local workspace", "main", "n/a", "2026-07-20", "Active", "KB assessments", "Markdown Python", "Documentation", "N/A", "N/A", "N/A", "References qTest", "N/A", "N/A", "N/A", "N/A", "N/A"],
        ["universal-platform-coverage", r"C:\Workspace\GitLab\qa-automation-kb\universal-platform-coverage", "local KB", "main", "2026-07-01 SME", "2026-07-01", "Reference", "UP assessment", "Assessment", "qTest/TestNG reconciliation", "CSV mappings", "N/A", "PDF/CSV", "744 qTest population", "API ops", "N/A", "N/A", "N/A", "N/A"],
    ]
    write_csv(
        "04-repositories/repository-inventory.csv",
        ["repository_name", "local_path", "remote_url", "current_branch", "head_commit", "last_relevant_commit_date", "status", "business_area", "framework_language", "test_types", "suite_files", "maven_gradle_profiles", "report_formats", "qtest_jira_identifiers", "endpoint_service_identifiers", "disabled_ignored_tests", "test_data_dependencies", "supported_environments", "idp_non_idp_support"],
        rows,
    )


def build_automation_test_inventory() -> None:
    rows = [
        ["T-001", "V3 UE regression", "prime-test-automation", "unite/unite-universal-enrollment", "TestNG", "303", "1", "Active recently executed", "GitLab scheduled_regression_job", "Hard scheduled gate", "Verified", "2026-07-20", "UP ledger"],
        ["T-002", "V3 Unite master", "prime-test-automation", "unite/unite", "TestNG", "133", "1", "Active recently executed", "GitLab scheduled_regression_job", "Hard scheduled gate", "Verified", "2026-07-20", "UP ledger 379 scoped"],
        ["T-003", "Metadataweb API", "api-test-automation", "jsonapi-metadataweb", "TestNG", "13 suites", "1", "Active recently executed", "scheduled_metadataweb_stage1", "Hard scheduled gate", "Verified", "2026-07-20", "gitlab-ci.yml"],
        ["T-004", "Mobile2 API master", "api-test-automation", "mobile/mobile2", "TestNG", "20 classes 24 endpoints", "1", "Active manually executed", "None scheduled", "Unverified CI", "Verified code", "2026-07-20", "QA-1405 pending"],
        ["T-005", "Mobile1 API", "api-test-automation", "mobile/mobile1", "TestNG", "6 classes", "1", "Active manually executed", "None", "Unverified CI", "Verified", "2026-07-20", "6/27 endpoints"],
        ["T-006", "V2 UI PRIME", "unite-test-automation", "unite/testsuite", "Cucumber", "2176 scenarios", "1", "Active manually executed", "Jenkins Ant inferred", "Soft/manual", "Inferred", "2026-07-20", "qTest 744 executed subset"],
        ["T-007", "V2 backoffice", "unite-test-automation", "backoffice", "Cucumber", "1077", "1", "Active manually executed", "Ant daily target", "Soft/manual", "Inferred", "2026-07-20", "Not on V3 schedule"],
        ["T-008", "ASTRO UI", "astro-test-automation", "astro/testsuite", "Cucumber", "1236", "1", "Active manually executed", "Jenkins refresh only", "Unverified", "Inferred", "2026-07-20", "No nightly GitLab"],
        ["T-009", "UP IDP perf", "performance-test-automation", "universal-platform/idp", "JMeter/Taurus", "6+", "1", "Integrated scheduled regression", "AGSUP_IDP_REGRESSION_SUITE", "Perf not functional gate", "Inferred", "2026-07-02", "KB tracker"],
        ["T-010", "MSC service BDD", "UniteMSC", "*/src/test/resources/features", "Cucumber", "112 aggregate", "14", "Integrated deployment validation", "Per-repo GitLab", "Build/test gate", "Verified", "2026-07-20", "Not BFF acceptance"],
        ["T-011", "Universal API modules", "api-test-automation", "universal/*", "TestNG", "38 classes", "1", "Code exists unverified schedule", "None", "Unverified CI", "Verified code", "2026-07-20", "Manual only"],
        ["T-012", "Mobile2 banks smoke", "api-test-automation", "banks-smoke-testng.xml", "TestNG", "2", "1", "Disabled from master by design", "Smoke manual", "Informational", "Verified", "2026-07-20", "Excluded from master"],
        ["T-013", "GitHub Mobile2 Nexus", "api-test-automation", "documented only", "N/A", "0", "1", "Planned", "None", "Planned", "Planned", "2026-07-20", "No workflow file"],
        ["T-014", "COPACS", "N/A", "N/A", "N/A", "0", "0", "Unknown", "None", "None", "Unknown", "2026-07-20", "No repo identified"],
    ]
    write_csv(
        "04-repositories/automation-test-inventory.csv",
        ["asset_id", "name", "repository", "path", "framework", "count_unit", "count", "classification", "ci_integration", "gate_type", "status", "evidence_date", "notes"],
        rows,
    )


def build_identifier_mapping() -> None:
    rows = [
        ["ID-001", "Jira issue key", "ACS-I-2679", "qTest module", "PROVISIONAL-MOD-01", "Inferred", "Medium", "Module aggregate 151 cases", "SME confirmed UP scope"],
        ["ID-002", "Jenkins target", "stage1-enrollments-regression", "Suite XML", "stage1-enrollments.xml", "Verified", "High", "Jenkins log extract", "unite-test-automation"],
        ["ID-003", "qTest cycle", "PRIME CL-*", "Test run", "TR-*", "Inferred", "Medium", "PDF report format", "Spike docs reference"],
        ["ID-004", "Endpoint", "GET /mobile/dashboard", "Test class", "MobileDashboardRequestTest", "Verified", "High", "Code @ cee0de9", "api-test-automation"],
        ["ID-005", "Maven profile", "stage1-unite-regression-master", "Pipeline job", "scheduled_regression_job", "Verified", "High", "prime-test-automation CI", "GitLab"],
        ["ID-006", "qTest case ID", "Row-level", "Repository test", "N/A", "Unknown", "Low", "Not exported", "Blocked without qTest API"],
    ]
    write_csv(
        "04-repositories/identifier-mapping-candidates.csv",
        ["mapping_id", "identifier_type", "source_value", "target_type", "target_value", "match_status", "confidence", "notes", "evidence"],
        rows,
    )


def build_pipeline_inventory() -> None:
    rows = [
        ["GitLab", "scheduled_regression_job", "prime-test-automation", "schedules", "Nightly inferred", "Stage1", "mvn test UE + Unite profiles", "stage1-ue-regression-test; stage1-unite-regression-master", "V3 UI", "JUnit surefire", "Always", "Email via GitLab default", "exit 1 on fail", "Blocking scheduled", "Pre-release signal not MR gate", "Not live-verified", "Unknown", "Hard scheduled regression", "prime-test-automation/.gitlab-ci.yml", "Verified YAML"],
        ["GitLab", "scheduled_metadataweb_stage1", "api-test-automation", "schedules", "Nightly ~1AM comment", "Stage1", "mvn test metadataweb", "stage1-api-metadata-pipeline", "Universal metadata API", "JUnit", "Always", "Unknown", "Maven fail", "Blocking scheduled", "Partial UP API", "Not live-verified", "Unknown", "Hard scheduled regression", "api-test-automation/.gitlab-ci.yml", "Verified YAML"],
        ["GitLab", "stage5_smoke_job", "prime-test-automation", "manual schedule", "Manual", "Stage5", "stage5 smoke profiles", "stage5-ue-smoke-test", "V3 smoke", "JUnit", "Always", "Unknown", "exit 1", "Manual smoke", "Deployment prep", "On demand", "Unknown", "Manual smoke", "prime-test-automation/.gitlab-ci.yml", "Verified"],
        ["GitLab", "unitemsc pipeline", "UniteMSC services", "push/MR", "Per commit", "Dev/test", "mvn test", "default", "Microservice", "Jacoco Cucumber", "Template-defined", "Unknown", "Maven fail typical", "Build/deploy gate", "Service not GS E2E", "Per commit", "Unknown", "Service CI gate", "unite-mobile1/.gitlab-ci.yml", "Verified template ref"],
        ["GitHub Actions", "Mobile2 Nexus", "api-test-automation", "N/A", "Planned", "QC4", "Documented workflow", "mobile-ms-nexus-ci", "Mobile2 deploy validation", "N/A", "N/A", "N/A", "N/A", "Planned", "N/A", "N/A", "N/A", "Planned", "mobile/project-documents/17-*.md", "Planned"],
        ["Jenkins", "AGSUP_IDP_REGRESSION_SUITE", "performance-test-automation", "cron weekdays", "H 3 * * 1-5", "Load servers", "Taurus YAML", "IDP perf suite", "UP performance", "Perf reports", "Unknown", "Unknown", "Unknown", "Scheduled perf", "Not functional gate", "2026-07-02 doc", "Unknown", "Perf regression", "mobile2-api-db-validation tracker", "Inferred"],
        ["Jenkins", "V2 UI regression", "unite-test-automation", "Unknown", "Unknown", "Stage1", "Ant targets", "regression-frontoffice-daily", "V2 UI", "TestNG HTML", "Unknown", "Unknown", "Unknown", "Soft/manual", "Unknown", "Stale", "Unknown", "Legacy", "build.xml + qTest PDF", "Inferred"],
        ["Manual", "Mobile2 API local", "api-test-automation", "Developer", "Ad hoc", "QC4/Stage1", "mvn -P profiles", "mobile2 suites", "MSC API", "Surefire", "Local", "N/A", "N/A", "None", "N/A", "2026-07-20", "Pass/fail local", "Unverified schedule", "Local execution", "Verified"],
    ]
    write_csv(
        "05-pipelines/pipeline-job-inventory.csv",
        ["platform", "job_or_workflow_name", "repository", "trigger", "schedule", "environment", "test_command", "suite_or_profile", "business_scope", "report_artifact", "report_retention", "notifications", "allow_failure", "blocking_behavior", "deployment_relationship", "latest_available_run", "latest_result", "ci_classification", "evidence_source", "status"],
        rows,
    )


def build_source_code_coverage_audit() -> None:
    rows = [
        ["unite-mobile2", "Application microservice", "Yes", "Unit + Cucumber", "Jacoco maven plugin 0.8.5", "CI report stage", "GitLab job artifacts", "No threshold verified", "No", "No merge block on coverage %", "Unknown", "Unknown", "Service tests only", "Inferred", "pom.xml + .gitlab-ci.yml"],
        ["unite-mobile1", "Application microservice", "Yes", "Unit + Cucumber", "Jacoco in pom", "CI report stage", "GitLab", "No threshold", "No", "No", "Unknown", "Unknown", "Service tests", "Inferred", "pom.xml"],
        ["api-test-automation", "QA automation repo", "No", "N/A", "No JaCoCo on test code", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "Black-box tests", "Verified", "Repo scan — QA not app coverage"],
        ["prime-test-automation", "QA automation repo", "No", "N/A", "No app source coverage", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "UI tests", "Verified", "Repo scan"],
        ["prime-test-automation/jsonapi libs", "Test libraries", "Partial", "Unit", "Jacoco refs in pom", "Local/CI unknown", "Unknown", "Unknown", "No", "No", "Unknown", "Unknown", "Library unit tests", "Inferred", "jsonapi-core pom hits"],
        ["UniteMSC aggregate", "14 services", "Yes", "Service unit", "Jacoco + RUN_SONARQUBE:false", "Template", "Not published to Sonar", "No GS-wide threshold", "No", "Build fail on test not coverage", "Unknown", "Unknown", "Unit/service", "Verified", "CI variables"],
        ["Government Savings QA pipelines", "Regression pipelines", "No", "N/A", "No coverage_report in GitLab CI reviewed", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "Business test pass % only", "Verified", "gitlab-ci.yml review"],
    ]
    write_csv(
        "05-pipelines/source-code-coverage-audit.csv",
        ["repository", "repo_type", "coverage_generated", "test_type_generating", "tooling", "report_uploaded", "visible_in_platform", "minimum_threshold", "decrease_blocks_approval", "failure_blocks_merge_deploy", "branch_coverage", "report_current", "integration_tests_included", "status", "evidence"],
        rows,
    )


def build_cross_system_traceability() -> None:
    rows = [
        ["R-001", "ACS-I-2679", "PROVISIONAL-MOD-01", "151 qTest cases", "stage1-enrollments.xml", "stage1-enrollments-regression", "Jenkins inferred", "Stale/unknown", "N/A", "Partial", "Module map", "Medium", "2026-06-29", "Refresh qTest row IDs"],
        ["R-002", "N/A", "N/A", "N/A", "MobileDashboardRequestTest", "N/A", "Manual/local", "Unknown", "N/A", "Unmatched", "Endpoint class only", "High", "2026-07-20", "Add qTest + Jira links"],
        ["R-003", "N/A", "V3 TestNG 379", "N/A", "prime UE+Unite tests", "scheduled_regression_job", "GitLab", "Unknown live", "N/A", "Partial", "UP ledger not qTest", "High", "2026-07-01", "V3 not qTest-keyed"],
        ["R-004", "QA-1405", "Planned", "N/A", "mobile2 master suite", "Planned GitLab job", "None", "N/A", "N/A", "Planned", "Story to pipeline", "High", "Planned", "Implement + link"],
        ["R-005", "ACS-I-2680", "PROVISIONAL-MOD-02", "27 cases", "stage1-web-login.xml", "Jenkins", "Stale", "N/A", "Partial", "Module map", "Medium", "2026-06-29", "SME verified IDP scope"],
    ]
    write_csv(
        "07-reconciliation/cross-system-traceability.csv",
        ["record_id", "jira_key", "qtest_ref", "qtest_count", "repo_test", "suite_profile", "pipeline_job", "latest_execution", "source_code_coverage", "match_status", "match_method", "confidence", "data_freshness", "remediation"],
        rows,
    )


def build_unmatched_records() -> None:
    rows = [
        ["U-001", "Jira scope", "ACS-I-2682 Angular", "No qTest module in UP 744 map", "Inferred", "Add scope rule or map to Angular tests"],
        ["U-002", "qTest", "179 PRIME expansion", "No Jenkins module breakdown", "Inferred", "SME allocate to workstreams"],
        ["U-003", "Repository", "92 api-test-automation tests", "No qTest IDs in source", "Verified", "Automation ID standard"],
        ["U-004", "Repository", "Mobile2 20 classes", "No scheduled pipeline", "Verified", "QA-1405"],
        ["U-005", "Pipeline", "Universal API modules", "No GitLab schedule", "Verified", "Expand CI schedule"],
        ["U-006", "qTest", "744 V2 population", "V3 GitLab uses TestNG not qTest", "Verified", "Separate denominators per framework"],
    ]
    write_csv(
        "07-reconciliation/unmatched-records.csv",
        ["unmatched_id", "system", "record", "reason", "status", "remediation"],
        rows,
    )


def build_data_conflicts() -> None:
    rows = [
        ["C-001", "Metric definition", "Leadership '80% covered'", "Multiple incompatible metrics exist", "Verified", "Never combine without formula", "Use separate A–E metrics"],
        ["C-002", "Mobile2 coverage", "88% leadership Jul 14 vs 100% Jul 20", "Denominator/coverage definition changed", "Verified", "Superseded by code review", "Use 24/24 in-scope endpoints"],
        ["C-003", "V2 execution", "qTest 744 executed vs Jenkins 508 subset", "Different counting units", "Verified", "Document subset relationship", "UP reconciliation ledger"],
        ["C-004", "GitLab schedule", "KB schedule #3961313 vs live unverified", "No API access", "Unknown", "Live-verify schedule", "DevOps screenshot"],
    ]
    write_csv(
        "07-reconciliation/data-conflicts.csv",
        ["conflict_id", "domain", "value_a", "value_b", "status", "resolution", "evidence"],
        rows,
    )


def build_denominator_xlsx() -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("openpyxl not installed — skipping denominator-decision-matrix.xlsx")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Denominator Options"
    headers = [
        "Model", "Measures", "Data source", "Completeness", "Data quality",
        "Maintenance burden", "Manipulation risk", "Supports automation",
        "Leadership clarity", "Reproducible", "Recommended use", "Status",
    ]
    ws.append(headers)
    data = [
        ["Approved business capabilities", "Scope items leadership agrees are in-scope", "Jira epics + capability map", "Low", "Medium", "High", "Medium", "Yes", "High", "Medium", "Enterprise scope denominator", "Planned"],
        ["Jira acceptance criteria", "Testable AC per story", "Jira API", "Unknown", "Unknown", "Medium", "High", "Partial", "Medium", "Unknown", "Story-level API/UI", "Blocked"],
        ["qTest approved regression", "Approved automated scenarios", "qTest API", "Partial", "Stale export", "Medium", "Medium", "Yes", "High", "Medium", "UI regression denominator", "Partial"],
        ["OpenAPI endpoints", "HTTP operations in scope", "Service repos + registries", "Partial", "Medium", "Medium", "Low", "Yes", "High", "High", "API automation denominator", "Verified subset"],
        ["Service catalog", "Microservices in scope", "Architecture docs", "Low", "Unknown", "High", "Low", "Partial", "Medium", "Low", "Integration scope", "Planned"],
        ["Critical user journeys", "End-to-end flows", "BA + qTest", "Medium", "Medium", "High", "Medium", "Yes", "High", "Medium", "Smoke/deployment gates", "Inferred"],
        ["Risk-based controls", "High-risk scenarios", "Risk register", "Low", "Unknown", "Medium", "Low", "Yes", "High", "Medium", "Gate coverage numerator", "Planned"],
        ["JaCoCo lines/branches", "Application code exercised", "Service CI + Sonar", "Partial", "Medium", "Low", "Low", "Yes", "Medium", "High", "Source-code coverage (A)", "Partial"],
    ]
    for row in data:
        ws.append(row)
    fill = PatternFill("solid", fgColor="003057")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
    out = ROOT / "06-model" / "denominator-decision-matrix.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote {out}")


def main() -> None:
    build_integration_matrix()
    build_qtest_data_quality()
    build_jira_link_quality()
    build_repository_inventory()
    build_automation_test_inventory()
    build_identifier_mapping()
    build_pipeline_inventory()
    build_source_code_coverage_audit()
    build_cross_system_traceability()
    build_unmatched_records()
    build_data_conflicts()
    build_denominator_xlsx()
    print(f"CSV artifacts generated under {ROOT}")


if __name__ == "__main__":
    main()
